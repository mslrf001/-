import re
import pandas as pd
import datetime
import sys
import tkinter as tk
from tkinter import scrolledtext, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from pypinyin import lazy_pinyin
from difflib import SequenceMatcher
import threading
import json
import argparse
from pathlib import Path
from Levenshtein import distance as levenshtein_distance

# 统一数据处理器
class DataProcessor:
    """数据处理工具类"""
    
    CHINESE_NUM_DICT = {
        '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10, '两': 2,
        '壹': 1, '贰': 2, '叁': 3, '肆': 4, '伍': 5, '陆': 6, '柒': 7, '捌': 8, '玖': 9, '拾': 10
    }
    
    @staticmethod
    def to_pinyin(text):
        """将中文转换为拼音"""
        return ''.join(lazy_pinyin(text))
    
    @staticmethod
    def pinyin_similarity(name1, name2):
        """计算拼音相似度"""
        if not name1 or not name2:
            return 0.0
        pinyin1 = DataProcessor.to_pinyin(name1)
        pinyin2 = DataProcessor.to_pinyin(name2)
        return SequenceMatcher(None, pinyin1, pinyin2).ratio()
    
    @staticmethod
    def find_manager(branch, text, branch_managers):
        """查找经理"""
        managers = branch_managers.get(branch, [])
        
        # 先尝试精确匹配经理姓名
        for manager in managers:
            if manager in text:
                return manager
        
        # 提取所有中文姓名（2-4个汉字）
        chinese_words = re.findall(r'[\u4e00-\u9fa5]{2,4}', text)
        
        # 优先匹配经理姓名
        for manager in managers:
            for word in chinese_words:
                if manager == word:
                    return manager
        
        # 使用拼音模糊匹配，降低阈值
        for manager in managers:
            for word in chinese_words:
                similarity = DataProcessor.pinyin_similarity(manager, word)
                if similarity > 0.6:  # 降低相似度阈值
                    return manager
        
        # 检查文本中是否包含经理姓名的部分匹配
        for manager in managers:
            for word in chinese_words:
                if len(manager) >= 2 and len(word) >= 2:
                    # 检查是否包含经理姓名中的字
                    if manager[0] in word and manager[1] in word:
                        return manager
        
        return None
    
    @staticmethod
    def extract_number(text):
        """提取中文或数字"""
        if not text:
            return 1
        
        text_clean = text.replace(' ', '')
        patterns = [
            r'(\d+|[一|二|三|四|五|六|七|八|九|十|两|壹|贰|叁|肆|伍|陆|柒|捌|玖|拾])\s*户',
            r'(\d+|[一|二|三|四|五|六|七|八|九|十|两|壹|贰|叁|肆|伍|陆|柒|捌|玖|拾])\s*户.*接入号'
        ]
        
        for pattern_str in patterns:
            match = re.search(pattern_str, text_clean)
            if match:
                num_str = match.group(1)
                if num_str.isdigit():
                    return int(num_str)
                else:
                    return DataProcessor.CHINESE_NUM_DICT.get(num_str, 1)
        
        return 1
    
    # 预编译的正则表达式作为类变量
    _POINTS_PATTERNS = [
        re.compile(r'(?:得|加|\+|增加)?\s*(\d+)\s*(?:分|积分)?'),
        re.compile(r'(\d+)顺档(\d+)\+(\d+)'),  # X顺档Y+Z元
        re.compile(r'顺档\d*\+(\d+)'),         # 顺档Y+Z元
        re.compile(r'(\d+)顺档\+(\d+)')        # X顺档+Z元
    ]
    
    @staticmethod
    def extract_points_optimized(text, points_regex=None):
        """优化的积分提取方法，统一处理常规积分和顺档格式"""
        # 使用自定义正则或默认正则
        if points_regex:
            custom_pattern = re.compile(points_regex)
            match = custom_pattern.search(text)
            if match:
                try:
                    return int(match.group(1))
                except (ValueError, IndexError):
                    pass
        
        # 常规积分提取
        match = DataProcessor._POINTS_PATTERNS[0].search(text)
        if match:
            try:
                return int(match.group(1))
            except (ValueError, IndexError):
                pass
        
        # 顺档格式处理
        for pattern in DataProcessor._POINTS_PATTERNS[1:]:
            match = pattern.search(text)
            if match:
                groups = match.groups()
                try:
                    if len(groups) == 3:  # X顺档Y+Z元
                        return int(groups[1]) - int(groups[0]) + int(groups[-1])
                    else:  # 其他格式
                        return int(groups[-1])
                except (ValueError, IndexError):
                    continue
        
        return None

# 配置管理器
class ConfigManager:
    """统一的配置管理器"""
    def __init__(self, config_dir=None):
        self.config_dir = config_dir or ""
        self._business_config = None
        self._branch_managers = None
        self._branch_channel_map = None
        self._loaded = False
    
    def _ensure_loaded(self):
        """确保配置已加载（延迟加载）"""
        if not self._loaded:
            self._load_all_configs()
            self._loaded = True
    
    def _load_all_configs(self):
        """加载所有配置文件"""
        self._business_config = self._load_json_config('存量业务配置.json')
        self._branch_managers = self._load_manager_config()
        self._branch_channel_map = self._load_channel_config()
    
    @property
    def business_config(self):
        self._ensure_loaded()
        return self._business_config
    
    @property
    def branch_managers(self):
        self._ensure_loaded()
        return self._branch_managers
    
    @property
    def branch_channel_map(self):
        self._ensure_loaded()
        return self._branch_channel_map
    
    def _load_json_config(self, filename):
        """加载JSON配置文件"""
        config_path = Path(self.config_dir) / filename if self.config_dir else Path(filename)
        
        # 尝试多个路径
        paths_to_try = [
            config_path,
            Path("config") / filename,
            Path(filename)
        ]
        
        for path in paths_to_try:
            try:
                if path.exists():
                    with open(path, 'r', encoding='utf-8') as f:
                        return json.load(f)
            except (FileNotFoundError, json.JSONDecodeError):
                continue
        return None
    
    def _load_manager_config(self):
        """加载存量经理配置"""
        manager_config = self._load_json_config('存量经理配置.json')
        if manager_config:
            return manager_config.get("business_categories", {}).get("branch_managers", {}).get("branch_manager_map", {})
        return {}

    def _load_channel_config(self):
        """加载渠道厅店配置"""
        channel_config = self._load_json_config('渠道厅店配置.json')
        if channel_config:
            return channel_config.get("business_categories", {}).get("channel_stores", {}).get("branch_channel_map", {})
        return {}

# 全局配置管理器（延迟初始化）
config_manager = None

def get_config_manager():
    """获取配置管理器实例（延迟初始化）"""
    global config_manager
    if config_manager is None:
        config_manager = ConfigManager()
    return config_manager

def generate_report(dragon_text, status_callback, config_dir=None):
    """
    根据输入的接龙文本生成统计报表

    参数:
        dragon_text: 包含【锁存】和【欠停】部分的接龙文本
        status_callback: 用于更新GUI状态的回调函数
        config_dir: 配置文件目录路径（可选）
    返回:
        生成的报表文件名或None
    """
    try:
        # 获取当前日期时间用于文件名和标题
        file_datetime = datetime.datetime.now().strftime("%m%d_%H%M")  # 文件名格式：月日_时分
        title_datetime = datetime.datetime.now().strftime("%m%d_%H:%M")  # 标题格式：月日_时:分

        # 使用ConfigManager加载配置文件
        config_manager = get_config_manager()

        # 初始化所有存量经理的数据（固定数据，即使无业务也显示）
        manager_data = {}  # (branch, manager) -> dict
        
        # 统一数据处理器
        class ReportProcessor:
            """报表数据处理器"""
            
            def __init__(self, branch_managers, business_config):
                self.branch_managers = branch_managers
                self.business_config = business_config
                self.manager_data = {}
                self.exception_records = []
                self.processed_lines = set()  # 记录已处理的记录
                self.all_lines = set()  # 记录所有记录
                self._category_configs = {}  # 缓存配置数据
                self._branch_pattern = re.compile(r'[支局分局]')  # 预编译分支局模式
                self._digit_pattern = re.compile(r'\d+')  # 预编译数字模式
                self._init_manager_data()
            
            def _init_manager_data(self):
                """初始化所有经理数据"""
                for branch, managers in self.branch_managers.items():
                    for manager in managers:
                        key = (branch, manager)
                        self.manager_data[key] = {
                            "分支局": branch, 
                            "存量经理": manager, 
                            "晒照": "", 
                            "锁存": 0,
                        "当月复机": 0,
                        "上月复机": 0,
                        "高危复机": 0,
                        "拆机挽留": 0,
                        "降档挽留": 0,
                        "合计": 0
                        }
            
            def _get_config_keywords(self, category):
                """获取配置关键词"""
                if not self.business_config or "business_categories" not in self.business_config:
                    return None
                return self.business_config["business_categories"].get(category)
            
            def _should_exclude(self, text, exclude_keywords):
                """检查是否应该排除"""
                return any(keyword in text for keyword in exclude_keywords)
            
            def _match_keywords(self, text, keywords):
                """匹配关键词"""
                return any(keyword in text for keyword in keywords)
            
            def _extract_branch_and_manager(self, line):
                """提取分支局和经理"""
                # 获取所有分支局名称，支持简称匹配
                branches = list(self.branch_managers.keys())
                
                # 先尝试精确匹配分支局全称
                for branch in branches:
                    if branch in line:
                        remaining = re.sub(re.escape(branch), '', line)
                        manager = DataProcessor.find_manager(branch, remaining, self.branch_managers)
                        if manager:
                            return branch, manager, remaining
                
                # 尝试匹配分支局简称（去掉"支局"或"分局"后缀）
                for branch in branches:
                    branch_short = branch.replace('支局', '').replace('分局', '')
                    if branch_short in line:
                        remaining = re.sub(re.escape(branch_short), '', line)
                        manager = DataProcessor.find_manager(branch, remaining, self.branch_managers)
                        if manager:
                            return branch, manager, remaining
                
                # 使用模糊匹配
                for branch in branches:
                    branch_short = branch.replace('支局', '').replace('分局', '')
                    if branch_short in line or branch in line:
                        remaining = re.sub(re.escape(branch_short), '', line)
                        manager = DataProcessor.find_manager(branch, remaining, self.branch_managers)
                        if manager:
                            return branch, manager, remaining
                
                return None, None, None
            
            # 预编译标题行检测正则表达式
            _TITLE_PATTERNS = [
                re.compile(r'^\[.*?\]'),  # 以方括号开头
                re.compile(r'^【.*?】'),   # 以中文方括号开头
                re.compile(r'^.*月.*日'),  # 包含月日格式
                re.compile(r'^.*接龙.*群'),  # 包含接龙群
                re.compile(r'^.*循环.*服务'),  # 包含循环服务
                re.compile(r'^.*拆降.*挽留'),  # 包含拆降挽留
                re.compile(r'^(\d+|[一二三四五六七八九十壹贰叁肆伍陆柒捌玖拾])')  # 数字开头
            ]
            
            _TITLE_KEYWORDS = {'接龙', '循环', '服务', '挽留', '群', '统计'}
            
            def _is_title_line(self, line):
                """判断是否为标题行"""
                line = line.strip()
                if not line:
                    return True
                
                # 检查标题模式
                for pattern in self._TITLE_PATTERNS[:-1]:  # 除了最后一个模式
                    if pattern.match(line):
                        return True
                
                # 检查行首是否有数字
                if not self._TITLE_PATTERNS[-1].match(line.replace(' ', '')):
                    # 如果行首不是数字，检查特定关键词
                    return any(keyword in line for keyword in self._TITLE_KEYWORDS)
                
                return False

            def _update_total(self, key):
                """更新合计值"""
                data = self.manager_data[key]
                data["合计"] = sum([
                    data.get("锁存", 0),
                    data.get("当月复机", 0),
                    data.get("上月复机", 0),
                    data.get("高危复机", 0),
                    data.get("拆机挽留", 0),
                    data.get("降档挽留", 0)
                ])
            


            def process_all_data(self, lines):
                """处理所有业务数据 - 基于配置关键字智能匹配，不再依赖固定标识符"""
                # 业务分类到中文列名的映射（缓存为实例变量）
                if not hasattr(self, '_category_mapping'):
                    self._category_mapping = {
                        "lock_storage": "锁存",
                        "current_month_recovery": "当月复机",
                        "last_month_recovery": "上月复机",
                        "high_risk_recovery": "高危复机",
                        "dismantle_retention": "拆机挽留",
                        "downgrade_retention": "降档挽留"
                    }
                
                # 缓存配置数据
                if not self._category_configs:
                    for category in self._category_mapping:
                        self._category_configs[category] = self._get_config_keywords(category)
                
                # 预编译正则表达式
                branch_pattern = self._branch_pattern
                digit_pattern = self._digit_pattern
                
                for line in lines:
                    line = line.strip()
                    if not line or not digit_pattern.search(line):
                        continue
                    
                    # 检查是否为标题行 - 直接跳过，不记录为异常
                    if self._is_title_line(line):
                        continue

                    # 记录所有包含'支局'或'分局'的记录
                    if '支局' in line or '分局' in line:
                        self.all_lines.add(line)

                    # 提取分支局和经理
                    branch, manager, remaining = self._extract_branch_and_manager(line)
                    if not branch:
                        self.exception_records.append(line)
                        continue

                    key = (branch, manager)
                    if key not in self.manager_data:
                        continue

                    # 智能匹配业务类型
                    matched = False
                    
                    for category, chinese_name in self._category_mapping.items():
                        config = self._category_configs[category]
                        if config:
                            exclude_keywords = config.get("exclude_keywords", [])
                            keywords = config.get("keywords", [])
                            
                            # 检查排除关键字
                            if exclude_keywords and self._should_exclude(line, exclude_keywords):
                                continue
                            
                            # 匹配关键字
                            if keywords and self._match_keywords(line, keywords):
                                # 提取数量
                                if category in ["current_month_recovery", "last_month_recovery", "high_risk_recovery"]:
                                    num = DataProcessor.extract_number(line)
                                else:
                                    num = 1
                                
                                self.manager_data[key][chinese_name] = self.manager_data[key].get(chinese_name, 0) + num
                                self._update_total(key)
                                matched = True
                                break
                        else:
                            # 使用默认关键字匹配
                            default_keywords = self._get_default_keywords(category)
                            if default_keywords and any(keyword in line for keyword in default_keywords):
                                if category in ["current_month_recovery", "last_month_recovery", "high_risk_recovery"]:
                                    num = DataProcessor.extract_number(line)
                                else:
                                    num = 1
                                
                                self.manager_data[key][chinese_name] = self.manager_data[key].get(chinese_name, 0) + num
                                self._update_total(key)
                                matched = True
                                break
                    
                    if matched:
                        self.processed_lines.add(line)
                    else:
                        self.exception_records.append(line)

            def _get_default_keywords(self, category):
                """获取默认关键字 - 现在完全依赖配置文件"""
                return []
            
            def _process_unmatched_records(self):
                """处理所有未被匹配的记录"""
                # 找出所有未被处理的记录
                for line in self.all_lines:
                    if line.strip() not in self.processed_lines:
                        # 检查是否为有效记录
                        branch, manager, _ = self._extract_branch_and_manager(line)
                        if not branch or not manager:
                            # 分支局或经理不在配置中，视为异常
                            if line.strip() not in self.exception_records:
                                self.exception_records.append(line.strip())
                        else:
                            # 分支局和经理都在配置中，但业务类型不在配置中，也视为异常
                            if line.strip() not in self.exception_records:
                                self.exception_records.append(line.strip())
            
            def generate_report(self):
                """生成报表数据"""
                # 处理所有未被匹配的记录
                self._process_unmatched_records()
                
                # 预计算分支局分组和排序
                branch_groups = {}
                all_managers = []
                
                # 单次遍历构建分组和收集所有数据
                for (branch, manager), data in self.manager_data.items():
                    if branch not in branch_groups:
                        branch_groups[branch] = []
                    branch_groups[branch].append(data)
                    all_managers.append(data)
                
                # 预计算总计值
                total_values = {
                    "锁存": sum(m["锁存"] for m in all_managers),
                    "当月复机": sum(m["当月复机"] for m in all_managers),
                    "上月复机": sum(m["上月复机"] for m in all_managers),
                    "高危复机": sum(m["高危复机"] for m in all_managers),
                    "拆机挽留": sum(m["拆机挽留"] for m in all_managers),
                    "降档挽留": sum(m["降档挽留"] for m in all_managers),
                    "合计": sum(m["合计"] for m in all_managers)
                }
                
                # 构建最终报表数据
                report = []
                
                # 按分支局名称排序并处理
                for branch in sorted(branch_groups.keys()):
                    branch_data = branch_groups[branch]
                    branch_data_sorted = sorted(branch_data, key=lambda x: x["存量经理"])
                    
                    # 添加经理数据
                    report.extend(branch_data_sorted)
                    
                    # 预计算分支局小计
                    branch_total_values = {
                        "锁存": sum(m["锁存"] for m in branch_data),
                        "当月复机": sum(m["当月复机"] for m in branch_data),
                        "上月复机": sum(m["上月复机"] for m in branch_data),
                        "高危复机": sum(m["高危复机"] for m in branch_data),
                        "拆机挽留": sum(m["拆机挽留"] for m in branch_data),
                        "降档挽留": sum(m["降档挽留"] for m in branch_data),
                        "合计": sum(m["合计"] for m in branch_data)
                    }
                    
                    # 添加分支局小计行
                    report.append({
                        "分支局": branch,
                        "存量经理": "小计",
                        "晒照": "",
                        **branch_total_values
                    })
                
                # 添加总计行
                report.append({
                    "分支局": "总计",
                    "存量经理": "",
                    "晒照": "",
                    **total_values
                })
                
                return report, self.exception_records
                
        # 使用ConfigManager加载配置文件
        config_manager = get_config_manager()
        
        # 获取存量经理配置
        branch_managers = config_manager.branch_managers
        
        # 获取存量业务配置
        business_config = config_manager.business_config
        
        # 创建处理器并开始处理数据
        processor = ReportProcessor(branch_managers, business_config)
        
        # 使用基于关键字的统一处理方式，不再依赖固定标识符
        lines = [line.strip() for line in dragon_text.split('\n') if line.strip()]
        
        # 使用ReportProcessor处理所有业务数据（基于关键字智能识别）
        processor.process_all_data(lines)
        
        # 生成最终报表数据
        report, exception_records = processor.generate_report()

        report_file = f"存量经理接龙数据通报_{file_datetime}.xlsx"
        with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
            df_report = pd.DataFrame(report)
            df_report.to_excel(writer, index=False, sheet_name="通报", startrow=1)
            ws = writer.sheets["通报"]

            title_style = Font(name="微软雅黑", bold=True, size=14, color="FFFFFF")
            title_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            header_style = Font(name="微软雅黑", bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

            title = f"存量经理接龙数据通报_{title_datetime}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = title_style
            title_cell.fill = title_fill
            title_cell.alignment = center_alignment

            for col_num, column_title in enumerate(df_report.columns, 1):
                cell = ws.cell(row=2, column=col_num, value=column_title)
                cell.font = header_style
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment

            data_start_row = 3
            data_end_row = len(df_report) + 2
            default_row_height = 18
            ws.row_dimensions[1].height = default_row_height * 1.2
            for row_num in range(2, data_end_row + 1):
                ws.row_dimensions[row_num].height = default_row_height
            column_widths = [12, 12, 10, 10, 10, 10, 10, 10, 10, 10]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # 设置数据行样式
            for row_num in range(data_start_row, data_end_row + 1):
                for col_num in range(1, 11):  # 包含所有业务类型和合计列
                    cell = ws.cell(row_num, column=col_num)
                    cell.border = border
                    cell.alignment = center_alignment
                    cell.font = Font(name="微软雅黑", size=11)
                if row_num <= len(df_report) + 1:
                    current_row = df_report.iloc[row_num - data_start_row]
                    if current_row["存量经理"] == "小计":
                        # 小计行样式
                        for col_num in range(1, 11):  # 包含所有业务类型和合计列
                            cell = ws.cell(row_num, column=col_num)
                            try:
                                cell.font = Font(name="微软雅黑", bold=True, size=11)
                            except AttributeError:
                                # 跳过已合并的单元格
                                pass
                        for col_num in range(2, 11):  # 整行标记，不包括第一列
                            cell = ws.cell(row_num, column=col_num)
                            try:
                                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                            except AttributeError:
                                # 跳过已合并的单元格
                                pass
                    elif current_row["分支局"] != "总计":
                        # 经理数据行样式
                        # 根据合计值来判断颜色标记
                        total_value = current_row["合计"]
                        if total_value >= 3:
                            # 合计值大于等于3，标记为红色（整行标记，不包括第一列）
                            for col_num in range(2, 11):  # 从第2列开始，包含存量经理列和合计列
                                cell = ws.cell(row_num, column=col_num)
                                try:
                                    cell.fill = red_fill
                                except AttributeError:
                                    # 跳过已合并的单元格
                                    pass
                        elif total_value == 0:
                            # 合计值为0，标记为绿色（整行标记，不包括第一列）
                            for col_num in range(2, 11):  # 从第2列开始，包含存量经理列和合计列
                                cell = ws.cell(row_num, column=col_num)
                                try:
                                    cell.fill = green_fill
                                except AttributeError:
                                    # 跳过已合并的单元格
                                    pass

            # 合并相同分支局名称的单元格并居中显示（参照渠道厅店样式）
            start_row = data_start_row
            for branch, group in df_report[df_report['分支局'] != '总计'].groupby('分支局', sort=False):
                # 合并整个分支局范围，包括小计行
                # 每个分支局包含数据行 + 小计行，所以需要合并的范围是数据行
                end_row = start_row + len(group) - 1
                first_cell = ws.cell(row=start_row, column=1)
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
                first_cell.font = Font(name="微软雅黑", size=11, bold=True)
                first_cell.border = border  # 添加框线
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                # 为合并单元格范围内的所有单元格设置框线和对齐
                for row_idx in range(start_row, end_row + 1):
                    cell = ws.cell(row=row_idx, column=1)
                    try:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    except AttributeError:
                        pass
                start_row = end_row + 1

            total_row = data_end_row
            for col_num in range(1, 11):  # 包含所有业务类型和合计列
                cell = ws.cell(total_row, column=col_num)
                try:
                    cell.font = Font(name="微软雅黑", bold=True)
                    cell.border = border
                    cell.alignment = center_alignment
                except AttributeError:
                    # 跳过已合并的单元格
                    pass

            # 添加点评文字
            comment_start_row = data_end_row + 3
            
            # 收集需要表扬和提醒的存量经理
            praised_managers = []
            reminded_managers = []
            
            # 统计业务量达3笔及以上和未破0的存量经理数量
            high_volume_managers = 0
            zero_volume_managers = 0
            
            for _, row_data in df_report.iterrows():
                if row_data["分支局"] == "总计" or row_data["存量经理"] == "小计":
                    continue
                manager_name = row_data["存量经理"]
                total_value = row_data["合计"]
                if total_value >= 3:
                    praised_managers.append(manager_name)
                    high_volume_managers += 1
                elif total_value == 0:
                    reminded_managers.append(manager_name)
                    zero_volume_managers += 1
            
            # 写入表扬文字
            if praised_managers:
                praise_text = f"截止目前，业务量达3笔及以上的存量经理有{high_volume_managers}人：{', '.join(praised_managers)}，特此表扬！"
                # 计算预估行高：每80个字符约一行，基础高度18，每行增加18
                estimated_height = max(18, 18 + (len(praise_text) // 80) * 18)
                ws.merge_cells(start_row=comment_start_row, start_column=1, end_row=comment_start_row, end_column=10)
                praise_cell = ws.cell(row=comment_start_row, column=1)
                praise_cell.value = praise_text
                praise_cell.font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")
                praise_cell.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                ws.row_dimensions[comment_start_row].height = estimated_height
                comment_start_row += 2
            
            # 写入提醒文字
            if reminded_managers:
                remind_text = f"截止目前，业务量未破0的存量经理有{zero_volume_managers}人：{', '.join(reminded_managers)}，请加油哦！"
                # 计算预估行高：每80个字符约一行，基础高度18，每行增加18
                estimated_height = max(18, 18 + (len(remind_text) // 80) * 18)
                ws.merge_cells(start_row=comment_start_row, start_column=1, end_row=comment_start_row, end_column=10)
                remind_cell = ws.cell(row=comment_start_row, column=1)
                remind_cell.value = remind_text
                remind_cell.font = Font(name="微软雅黑", size=11, bold=True, color="008000")
                remind_cell.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                ws.row_dimensions[comment_start_row].height = estimated_height

            # 添加分支局统计sheet
            # 筛选出分支局小计数据
            branch_total_df = df_report[df_report['存量经理'] == '小计'].copy()
            
            # 收集分支局统计数据
            branch_data = []
            if '合计' in branch_total_df.columns and '分支局' in branch_total_df.columns:
                for _, row_data in branch_total_df.iterrows():
                    if str(row_data['分支局']) != '总计' and str(row_data['分支局']) != 'nan':
                        try:
                            total = int(row_data['合计']) if pd.notna(row_data['合计']) else 0
                            branch_data.append({
                                '分支局': str(row_data['分支局']),
                                '业务量': total
                            })
                        except (ValueError, TypeError):
                            continue
            
            # 按业务量降序排序
            branch_data.sort(key=lambda x: x['业务量'], reverse=True)
            
            # 创建分支局统计sheet
            ws_branch = writer.book.create_sheet("分支局统计")
            
            # 设置分支局统计sheet标题
            ws_branch.merge_cells('A1:C1')
            title_cell = ws_branch['A1']
            title_cell.value = f'分支局业务量统计_{title_datetime}'
            title_cell.font = Font(name="微软雅黑", size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # 设置表头（与统计报表格式一致）
            branch_headers = ['排名', '分支局', '业务量(合计)']
            for col_idx, header in enumerate(branch_headers, 1):
                cell = ws_branch.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                cell.border = border
            
            # 填充分支局数据
            for rank, data in enumerate(branch_data, 1):
                row_idx = rank + 3
                ws_branch.cell(row=row_idx, column=1, value=rank).alignment = Alignment(horizontal='center')
                ws_branch.cell(row=row_idx, column=2, value=data['分支局']).alignment = Alignment(horizontal='left')
                ws_branch.cell(row=row_idx, column=3, value=data['业务量']).alignment = Alignment(horizontal='center')
                
                # 设置边框和样式（与统计报表一致）
                for col_idx in range(1, 4):
                    cell = ws_branch.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    cell.font = Font(name="微软雅黑", size=11)
                    if col_idx == 1:  # 排名列居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx == 2:  # 分支局列左对齐
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:  # 业务量列居中
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 添加汇总数据行
            total_row = len(branch_data) + 4
            ws_branch.cell(row=total_row, column=1, value="汇总").alignment = Alignment(horizontal='center')
            ws_branch.cell(row=total_row, column=2, value="合计").alignment = Alignment(horizontal='center')
            total_business = sum(data['业务量'] for data in branch_data)
            ws_branch.cell(row=total_row, column=3, value=total_business).alignment = Alignment(horizontal='center')
            
            # 设置汇总行的样式
            for col_idx in range(1, 4):
                cell = ws_branch.cell(row=total_row, column=col_idx)
                cell.border = border
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # 设置列宽（与统计报表一致）
            ws_branch.column_dimensions['A'].width = 8
            ws_branch.column_dimensions['B'].width = 25
            ws_branch.column_dimensions['C'].width = 15
            
            # 创建分支局柱状图
            branch_chart = BarChart()
            branch_chart.title = "分支局业务量排名"
            branch_chart.y_axis.title = "业务量"
            branch_chart.x_axis.title = "分支局"
            
            # 设置数据范围（不包含汇总行）
            branch_data_rows = len(branch_data)
            branch_data_ref = Reference(ws_branch, min_col=3, min_row=3, max_row=3+branch_data_rows, max_col=3)
            branch_categories = Reference(ws_branch, min_col=2, min_row=4, max_row=3+branch_data_rows)
            
            branch_chart.add_data(branch_data_ref, titles_from_data=True)
            branch_chart.set_categories(branch_categories)
            branch_chart.height = 10
            branch_chart.width = 20
            
            # 添加图表到分支局统计sheet
            ws_branch.add_chart(branch_chart, "E3")

            # 添加分析sheet - 存量经理按合计排名
            # 筛选出经理数据（排除小计和总计行）
            total_rows = len(df_report)
            manager_df = df_report[(df_report['存量经理'] != '小计') & (df_report['分支局'] != '总计')].copy()
            filtered_rows = len(manager_df)
            
            # 收集分析数据，参照渠道厅店的实现方式
            analysis_data = []
            
            # 确保manager_df有正确的列
            if '合计' in manager_df.columns and '存量经理' in manager_df.columns and '分支局' in manager_df.columns:
                for _, row_data in manager_df.iterrows():
                    if row_data['分支局'] != '总计' and str(row_data['存量经理']) != '小计' and str(row_data['存量经理']) != 'nan':
                        try:
                            total = int(row_data['合计']) if pd.notna(row_data['合计']) else 0
                            analysis_data.append({
                                '存量经理': str(row_data['存量经理']),
                                '合计': total,
                                '分支局': str(row_data['分支局'])
                            })
                        except (ValueError, TypeError):
                            continue
            
            # 始终创建分析sheet
            # 按合计值降序排序
            analysis_data.sort(key=lambda x: x['合计'], reverse=True)
            
            # 创建分析sheet
            ws_analysis = writer.book.create_sheet("分析")
            
            # 设置分析sheet标题
            ws_analysis.merge_cells('A1:C1')
            title_cell = ws_analysis['A1']
            title_cell.value = f'存量经理业务量排名分析_{title_datetime}'
            title_cell.font = Font(name="微软雅黑", size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # 设置表头
            headers = ['排名', '存量经理', '业务量(合计)']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_analysis.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                cell.border = border
            
            # 填充数据
            for rank, data in enumerate(analysis_data, 1):
                row_idx = rank + 3
                ws_analysis.cell(row=row_idx, column=1, value=rank).alignment = Alignment(horizontal='center')
                ws_analysis.cell(row=row_idx, column=2, value=data['存量经理']).alignment = Alignment(horizontal='left')
                ws_analysis.cell(row=row_idx, column=3, value=data['合计']).alignment = Alignment(horizontal='center')
                
                # 设置边框
                for col_idx in range(1, 4):
                    cell = ws_analysis.cell(row=row_idx, column=col_idx)
                    cell.border = border
            
            # 添加汇总数据行
            total_row = len(analysis_data) + 4
            ws_analysis.cell(row=total_row, column=1, value="汇总").alignment = Alignment(horizontal='center')
            ws_analysis.cell(row=total_row, column=2, value="合计").alignment = Alignment(horizontal='center')
            total_business = sum(data['合计'] for data in analysis_data)
            ws_analysis.cell(row=total_row, column=3, value=total_business).alignment = Alignment(horizontal='center')
            
            # 设置汇总行的样式
            for col_idx in range(1, 4):
                cell = ws_analysis.cell(row=total_row, column=col_idx)
                cell.border = border
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # 设置列宽
            ws_analysis.column_dimensions['A'].width = 8
            ws_analysis.column_dimensions['B'].width = 40
            ws_analysis.column_dimensions['C'].width = 12
            
            # 创建柱状图
            chart = BarChart()
            chart.title = "存量经理业务量排名"
            chart.y_axis.title = "业务量"
            chart.x_axis.title = "存量经理"
            
            # 设置数据范围（不包含汇总行）
            data_rows = len(analysis_data)
            data = Reference(ws_analysis, min_col=3, min_row=3, max_row=3+data_rows, max_col=3)
            categories = Reference(ws_analysis, min_col=2, min_row=4, max_row=3+data_rows)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 10
            chart.width = 20
            
            # 添加图表到分析sheet
            ws_analysis.add_chart(chart, "E3")
            


        abnormal_text = "\n".join(exception_records)
        status_callback(f"报表生成成功！文件已保存为：{report_file}", "success", abnormal_text)
        return report_file, abnormal_text

    except Exception as e:
         status_callback(f"生成存量经理报表失败: {e}", "error", "")
         return None

def generate_new_business_report(dragon_text, status_callback, config_dir=None):
    """
    根据输入的接龙文本生成渠道厅店新增业务报表

    参数:
        dragon_text: 包含新增业务的接龙文本
        status_callback: 用于更新GUI状态的回调函数
    返回:
        生成的报表文件名或None
    """
    try:
        # 使用配置管理器加载配置文件
        if config_dir:
            local_config_manager = ConfigManager(config_dir)
            channel_config = local_config_manager._load_json_config('渠道厅店配置.json')
            business_config = local_config_manager._load_json_config('存量业务配置.json')
            manager_config = local_config_manager._load_json_config('存量经理配置.json')
        else:
            cm = get_config_manager()
            channel_config = cm._load_json_config('渠道厅店配置.json')
            business_config = cm._load_json_config('存量业务配置.json')
            manager_config = cm._load_json_config('存量经理配置.json')
            
        if not channel_config or 'branch_channel_map' not in channel_config:
            status_callback("错误：无法加载或解析 '渠道厅店配置.json'。", "red")
            return
        
        branch_channel_map = channel_config.get('branch_channel_map', {})
        points_regex_config = channel_config.get('points_regex', r'积分\s*(\d+)')
        if isinstance(points_regex_config, list):
            points_regex = '|'.join(points_regex_config)
        else:
            points_regex = points_regex_config

        file_datetime = datetime.datetime.now().strftime("%m%d_%H%M")  # 文件名格式：月日_时分
        title_datetime = datetime.datetime.now().strftime("%m%d_%H:%M")  # 标题格式：月日_时:分
        report_data = []
        for branch, channels in branch_channel_map.items():
            for channel_info in channels:
                report_data.append({"分支局": branch, "渠道名称": channel_info['name'], "业务笔数": 0, "积分": 0})

        abnormal_entries = []  # 用于存储异常条目
        total_entries = 0
        processed_entries = 0


        def enhanced_fuzzy_match(text, keywords):
            """增强的模糊匹配算法，支持拼音匹配和多种匹配策略"""
            text_lower = text.lower().strip()
            best_match = None
            best_score = 0.0
            best_method = ""
            
            if not text or not keywords:
                return None, 0.0, ""



            # 不过滤关键字，保留所有有效关键字
            filtered_keywords = keywords
            
            if not filtered_keywords:
                return None, 0.0, ""

            # 第一阶段：精确匹配
            for keyword in filtered_keywords:
                keyword_lower = keyword.lower().strip()
                if keyword_lower in text_lower:
                    # 根据关键字长度调整分数，降低短关键字的分数要求
                    if len(keyword_lower) <= 2:
                        score = 0.5  # 降低2字关键字的分数要求
                    elif len(keyword_lower) <= 4:
                        score = 0.6  # 降低4字关键字的分数要求
                    else:
                        score = 0.7  # 降低长关键字的分数要求
                    
                    if score > best_score:
                        best_match = keyword
                        best_score = score
                        best_method = "精确匹配"
                    continue

            # 第二阶段：分支局/经理名称匹配（简化处理）
            for keyword in filtered_keywords:
                keyword_lower = keyword.lower().strip()
                
                # 分支局名称匹配
                if ("支局" in text_lower or "分局" in text_lower) and keyword_lower in text_lower:
                    score = 0.85
                    if score > best_score:
                        best_match = keyword
                        best_score = score
                        best_method = "分支局匹配"
                
                # 经理姓名匹配
                elif len(keyword_lower) <= 4 and keyword_lower in text_lower:
                    score = 0.8
                    if score > best_score:
                        best_match = keyword
                        best_score = score
                        best_method = "经理匹配"
            
            return best_match if best_score >= 0.4 else None, best_score, best_method

        def find_best_match(text, channel_config_data):
            """根据关键字或拼音在文本中查找最匹配的渠道名称，针对三个配置文件的关键字"""

            # 收集所有三个配置文件的关键字
            all_keywords = []
            keyword_to_mapping = {}
            
            # 1. 收集渠道厅店配置.json的关键字
            branch_channel_map = channel_config_data.get('branch_channel_map', {})
            for branch, channels in branch_channel_map.items():
                for channel in channels:
                    for keyword in channel.get('keywords', []):
                        all_keywords.append(keyword)
                        keyword_to_mapping[keyword] = channel['name']
            
            # 2. 收集存量业务配置.json的关键字（使用已加载的配置）
            if business_config:
                for business_type, config in business_config.get('business_categories', {}).items():
                    for keyword in config.get('keywords', []):
                        all_keywords.append(keyword)
                        keyword_to_mapping[keyword] = f"业务_{business_type}"
            
            # 3. 收集存量经理配置.json的关键字（分支局名称和经理姓名）
            if manager_config:
                branch_managers = manager_config.get("business_categories", {}).get("branch_managers", {}).get("branch_manager_map", {})
                keywords = manager_config.get("business_categories", {}).get("branch_managers", {}).get("keywords", [])
                
                # 添加所有关键字到列表
                for keyword in keywords:
                    all_keywords.append(keyword)
                    if keyword in branch_managers:
                        keyword_to_mapping[keyword] = f"分支局_{keyword}"
                    else:
                        # 经理姓名
                        keyword_to_mapping[keyword] = f"经理_{keyword}"

            # 使用增强的模糊匹配
            matched_keyword, similarity, method = enhanced_fuzzy_match(text, all_keywords)
            
            if matched_keyword:
                mapping_result = keyword_to_mapping[matched_keyword]
                # 只返回渠道厅店的匹配结果，过滤掉业务和经理匹配
                if not mapping_result.startswith("业务_") and not mapping_result.startswith("分支局_") and not mapping_result.startswith("经理_"):
                    return mapping_result

            return None

        entries = dragon_text.strip().split('\n')
        total_entries = len(entries)
        for original_entry in entries:
            if not original_entry.strip(): continue

            # 提取原始序号
            original_number_match = re.match(r'^(\d+[.、，,)\s]*)', original_entry)
            original_number = original_number_match.group(1) if original_number_match else ''
            entry = re.sub(r'^\d+[.、，,)\s]*', '', original_entry).strip()

            matched_channel = find_best_match(entry, channel_config)
            if not matched_channel:
                abnormal_entries.append(original_entry) # 记录未匹配的原始条目
                continue

            # 提取积分
            points = DataProcessor.extract_points_optimized(entry, points_regex)
            points_match = points is not None
            
            if not points_match:
                abnormal_entries.append(original_entry) # 记录原始条目

                continue

            processed_entries += 1

            for row in report_data:
                if row["渠道名称"] == matched_channel:
                    row["业务笔数"] += 1
                    row["积分"] += points
                    break

        # 创建DataFrame并计算合计
        df = pd.DataFrame(report_data)
        final_df_rows = []
        for branch, group in df.groupby('分支局', sort=False):
            final_df_rows.append(group)
            total_row = pd.DataFrame([{"分支局": branch, "渠道名称": "小计", "业务笔数": group['业务笔数'].sum(), "积分": group['积分'].sum()}])
            final_df_rows.append(total_row)
        
        final_df = pd.concat(final_df_rows, ignore_index=True)
        grand_total = pd.DataFrame([{"分支局": "总计", "渠道名称": "", "业务笔数": df['业务笔数'].sum(), "积分": df['积分'].sum()}])
        final_df = pd.concat([final_df, grand_total], ignore_index=True)

        report_file = f"渠道厅店接龙数据通报_{file_datetime}.xlsx"
        # 创建一个包含统计信息的新DataFrame
        summary_data = {
            '统计项': ['总条目数', '成功处理条目数', '异常条目数'],
            '数值': [total_entries, processed_entries, len(abnormal_entries)]
        }
        summary_df = pd.DataFrame(summary_data)

        # 将异常条目转换为DataFrame，保留原始序号
        abnormal_df = pd.DataFrame({
            '异常条目': abnormal_entries
        })

        # 创建分支局统计sheet的数据
        branch_summary_data = []
        for _, row_data in final_df.iterrows():
            if row_data['渠道名称'] == '小计':
                branch_summary_data.append({
                    '分支局': row_data['分支局'],
                    '业务量(合计)': int(row_data['业务笔数']),
                    '积分(合计)': int(row_data['积分'])
                })
        
        # 按积分降序排序
        branch_summary_data.sort(key=lambda x: x['积分(合计)'], reverse=True)

        with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name="通报", header=False, startrow=2)

            ws = writer.sheets["通报"]

            # 设置标题和表头
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = f'渠道厅店接龙数据通报_{title_datetime}'
            title_cell.font = Font(name="等线", size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            headers = ["分支局", "渠道名称", "业务笔数", "积分"]
            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=i, value=header)
                cell.font = Font(name="等线", size=11, bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # 样式和合并单元格
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row_idx, row_data in final_df.iterrows():
                # 确定整行背景色
                if row_data['分支局'] == '总计':
                    # 总计行：浅黄色
                    row_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                elif row_data['渠道名称'] == '小计':
                    # 小计行：浅灰色
                    row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                else:
                    # 普通数据行：根据积分值设置背景色
                    points = row_data['积分']
                    if points >= 100:
                        row_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # 红色
                    elif points == 0:
                        row_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 绿色
                    else:
                        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色
                
                # 应用整行样式（跳过第1列分支局列，从第2列开始）
                for col_idx in range(2, 5):
                    cell = ws.cell(row=row_idx + 3, column=col_idx)
                    try:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = row_fill
                        # 小计行和总计行字体加粗
                        if row_data['渠道名称'] == '小计' or row_data['分支局'] == '总计':
                            cell.font = Font(name="等线", size=11, bold=True)
                    except AttributeError:
                        # 跳过已合并的单元格
                        pass

            start_row = 3
            for branch, group in final_df.groupby('分支局', sort=False):
                end_row = start_row + len(group) - 1
                first_cell = ws.cell(row=start_row, column=1)
                first_cell.alignment = Alignment(horizontal='center', vertical='center')
                first_cell.font = Font(name="等线", size=11, bold=True)
                first_cell.border = thin_border  # 添加框线
                
                # 为总计行不合并单元格，直接设置样式
                if branch == '总计':
                    # 总计行不合并，只设置当前单元格样式
                    pass
                else:
                    # 普通分支局合并单元格
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    # 为合并单元格范围内的所有单元格设置框线和对齐
                    for row_idx in range(start_row, end_row + 1):
                        cell = ws.cell(row=row_idx, column=1)
                        try:
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        except AttributeError:
                            pass
                start_row = end_row + 1

            # 设置列宽和行高
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10

            # 创建分支局统计sheet（放在分析sheet之前）
            if branch_summary_data:
                branch_ws = writer.book.create_sheet("分支局统计")
                
                # 设置标题
                branch_ws['A1'] = '分支局积分排名统计_' + title_datetime
                branch_ws['A1'].font = Font(name="等线", size=14, bold=True)
                branch_ws['A1'].alignment = Alignment(horizontal='center')
                branch_ws.merge_cells('A1:C1')
                
                # 设置表头
                headers = ['排名', '分支局', '积分(合计)']
                for col_idx, header in enumerate(headers, 1):
                    cell = branch_ws.cell(row=3, column=col_idx)
                    cell.value = header
                    cell.font = Font(name="等线", size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    cell.border = thin_border
                
                # 填充数据
                for rank, data in enumerate(branch_summary_data, 1):
                    row_idx = rank + 3
                    branch_ws.cell(row=row_idx, column=1, value=rank).alignment = Alignment(horizontal='center')
                    branch_ws.cell(row=row_idx, column=2, value=data['分支局']).alignment = Alignment(horizontal='left')
                    branch_ws.cell(row=row_idx, column=3, value=data['积分(合计)']).alignment = Alignment(horizontal='center')
                    
                    # 设置边框
                    for col_idx in range(1, 4):
                        cell = branch_ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                
                # 添加汇总数据行
                total_row = len(branch_summary_data) + 4
                branch_ws.cell(row=total_row, column=1, value="汇总").alignment = Alignment(horizontal='center')
                branch_ws.cell(row=total_row, column=2, value="合计").alignment = Alignment(horizontal='center')
                total_points = sum(data['积分(合计)'] for data in branch_summary_data)
                branch_ws.cell(row=total_row, column=3, value=total_points).alignment = Alignment(horizontal='center')
                
                # 设置汇总行的样式
                for col_idx in range(1, 4):
                    cell = branch_ws.cell(row=total_row, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(name="等线", size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                # 设置列宽
                branch_ws.column_dimensions['A'].width = 8
                branch_ws.column_dimensions['B'].width = 20
                branch_ws.column_dimensions['C'].width = 15
                
                # 创建柱状图
                branch_chart = BarChart()
                branch_chart.title = "分支局积分排名"
                branch_chart.y_axis.title = "积分"
                branch_chart.x_axis.title = "分支局"
                
                # 设置数据范围（不包含汇总行）
                data_rows = len(branch_summary_data)
                branch_data = Reference(branch_ws, min_col=3, min_row=3, max_row=3+data_rows, max_col=3)
                branch_categories = Reference(branch_ws, min_col=2, min_row=4, max_row=3+data_rows)
                
                branch_chart.add_data(branch_data, titles_from_data=True)
                branch_chart.set_categories(branch_categories)
                branch_chart.height = 10
                branch_chart.width = 20
                
                # 添加图表到分支局统计sheet
                branch_ws.add_chart(branch_chart, "E3")

            # 创建分析sheet
            analysis_ws = writer.book.create_sheet("分析")
            
            # 获取需要分析的数据（排除小计和总计行）
            analysis_data = []
            total_rows = len(final_df)
            filtered_rows = 0
            
            for _, row_data in final_df.iterrows():
                if row_data['分支局'] != '总计' and row_data['渠道名称'] != '小计':
                    analysis_data.append({
                        '渠道名称': row_data['渠道名称'],
                        '积分': int(row_data['积分']),
                        '分支局': row_data['分支局']
                    })
                    filtered_rows += 1
            
            if analysis_data:
                # 按积分降序排序
                analysis_data.sort(key=lambda x: x['积分'], reverse=True)
                
                # 设置分析sheet标题
                analysis_ws['A1'] = '渠道厅店积分排名分析_' + title_datetime
                analysis_ws['A1'].font = Font(name="等线", size=14, bold=True)
                analysis_ws['A1'].alignment = Alignment(horizontal='center')
                analysis_ws.merge_cells('A1:C1')
                
                # 设置表头
                headers = ['排名', '渠道名称', '积分']
                for col_idx, header in enumerate(headers, 1):
                    cell = analysis_ws.cell(row=3, column=col_idx)
                    cell.value = header
                    cell.font = Font(name="等线", size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    cell.border = thin_border
                
                # 填充数据
                for rank, data in enumerate(analysis_data, 1):
                    row_idx = rank + 3
                    analysis_ws.cell(row=row_idx, column=1, value=rank).alignment = Alignment(horizontal='center')
                    analysis_ws.cell(row=row_idx, column=2, value=data['渠道名称']).alignment = Alignment(horizontal='left')
                    analysis_ws.cell(row=row_idx, column=3, value=data['积分']).alignment = Alignment(horizontal='center')
                    
                    # 设置边框
                    for col_idx in range(1, 4):
                        cell = analysis_ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                
                # 添加汇总数据行
                total_row = len(analysis_data) + 4
                analysis_ws.cell(row=total_row, column=1, value="汇总").alignment = Alignment(horizontal='center')
                analysis_ws.cell(row=total_row, column=2, value="合计").alignment = Alignment(horizontal='center')
                total_points = sum(data['积分'] for data in analysis_data)
                analysis_ws.cell(row=total_row, column=3, value=total_points).alignment = Alignment(horizontal='center')
                
                # 设置汇总行的样式
                for col_idx in range(1, 4):
                    cell = analysis_ws.cell(row=total_row, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(name="等线", size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                # 设置列宽
                analysis_ws.column_dimensions['A'].width = 8
                analysis_ws.column_dimensions['B'].width = 40
                analysis_ws.column_dimensions['C'].width = 12
                
                # 创建柱状图
                chart = BarChart()
                chart.title = "渠道厅店积分排名"
                chart.y_axis.title = "积分"
                chart.x_axis.title = "渠道名称"
                
                # 设置数据范围（不包含汇总行）
                data_rows = len(analysis_data)
                data = Reference(analysis_ws, min_col=3, min_row=3, max_row=3+data_rows, max_col=3)
                categories = Reference(analysis_ws, min_col=2, min_row=4, max_row=3+data_rows)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.height = 10
                chart.width = 20
                
                # 添加图表到分析sheet
                analysis_ws.add_chart(chart, "E3")

            # 添加点评文字
            comment_start_row = len(final_df) + 5
            
            # 收集需要表扬和提醒的渠道厅店
            praised_channels = []
            reminded_channels = []
            
            # 统计业务积分达100及以上和未破0的渠道厅店数量
            high_points_channels = 0
            zero_points_channels = 0
            
            for _, row_data in final_df.iterrows():
                if row_data['分支局'] == '总计' or row_data['渠道名称'] == '小计':
                    continue
                channel_name = row_data['渠道名称']
                points = row_data['积分']
                if points >= 100:
                    praised_channels.append(channel_name)
                    high_points_channels += 1
                elif points == 0:
                    reminded_channels.append(channel_name)
                    zero_points_channels += 1
            
            # 写入表扬文字
            if praised_channels:
                praise_text = f"截止目前，业务积分达100及以上的渠道厅店有{high_points_channels}家：{', '.join(praised_channels)}，特此表扬！"
                # 计算预估行高：每45个字符约一行，基础高度18，每行增加18
                estimated_height = max(18, 18 + (len(praise_text) // 45) * 18)
                ws.merge_cells(start_row=comment_start_row, start_column=1, end_row=comment_start_row, end_column=4)
                praise_cell = ws.cell(row=comment_start_row, column=1)
                praise_cell.value = praise_text
                praise_cell.font = Font(name="等线", size=11, bold=True, color="FF0000")
                praise_cell.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                ws.row_dimensions[comment_start_row].height = estimated_height
                comment_start_row += 2
            
            # 写入提醒文字
            if reminded_channels:
                remind_text = f"截止目前，业务积分未破0的渠道厅店有{zero_points_channels}家：{', '.join(reminded_channels)}，请加油哦！"
                # 计算预估行高：每45个字符约一行，基础高度18，每行增加18
                estimated_height = max(18, 18 + (len(remind_text) // 45) * 18)
                ws.merge_cells(start_row=comment_start_row, start_column=1, end_row=comment_start_row, end_column=4)
                remind_cell = ws.cell(row=comment_start_row, column=1)
                remind_cell.value = remind_text
                remind_cell.font = Font(name="等线", size=11, bold=True, color="008000")
                remind_cell.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
                ws.row_dimensions[comment_start_row].height = estimated_height

        # 返回异常记录文本
        abnormal_text = "\n".join(abnormal_entries)
        status_callback(f"报表生成成功！文件已保存为：{report_file}", "success", abnormal_text)
        return report_file

    except Exception as e:
        status_callback(f"生成渠道厅店报表失败: {e}", "error", "")
        return None

class ReportApp(ttk.Window):
    def __init__(self):
        super().__init__(themename="simplex", title="接龙数据报表生成器")
        self.geometry("900x800")
        self.center_window()
        self.config_dir = global_config_dir  # 设置配置目录
        self.create_widgets()

    def create_widgets(self):
        """创建GUI组件"""
        # 创建主框架
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=BOTH, expand=YES)

        # 创建标签页控件
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=BOTH, expand=YES)
        
        # 设置标签页选中时的背景色为浅蓝色
        style = ttk.Style()
        style.configure('TNotebook.Tab', background='lightgray')
        style.map('TNotebook.Tab',
                 background=[('selected', '#ADD8E6'),  # 浅蓝色
                            ('active', '#E6F3FF')])   # 鼠标悬停时的颜色

        # 第一个标签页：存量经理业务
        self.create_manager_report_tab(notebook)

        # 第二个标签页：渠道厅店业务
        self.create_new_business_tab(notebook)

    def create_manager_report_tab(self, notebook):
        """创建存量经理报表标签页"""
        jelong_tab = ttk.Frame(notebook, padding="10")
        notebook.add(jelong_tab, text="存量经理业务")

        # 控制框架（底部）
        control_frame = ttk.Frame(jelong_tab)
        control_frame.pack(side=BOTTOM, fill=X, pady=10)

        # 内容框架（顶部，填充剩余空间）
        content_frame = ttk.Frame(jelong_tab)
        content_frame.pack(side=TOP, fill=BOTH, expand=YES)

        input_frame = ttk.Labelframe(content_frame, text="请粘贴接龙数据", padding="5")
        input_frame.pack(fill=BOTH, expand=YES, padx=5, pady=5)
        self.text_input = scrolledtext.ScrolledText(input_frame, wrap=tk.WORD, width=80, height=12, font=("微软雅黑", 10))
        self.text_input.pack(fill=BOTH, expand=YES)
        
        # 创建右键菜单
        self.text_input_menu = tk.Menu(self.text_input, tearoff=0)
        self.text_input_menu.add_command(label="剪切", command=self.cut_text_input)
        self.text_input_menu.add_command(label="复制", command=self.copy_text_input)
        self.text_input_menu.add_command(label="粘贴", command=self.paste_text_input)
        self.text_input_menu.add_separator()
        self.text_input_menu.add_command(label="全选", command=self.select_all_text_input)
        
        # 绑定右键事件
        self.text_input.bind("<Button-3>", self.show_text_input_menu)

        abnormal_frame = ttk.Labelframe(content_frame, text="异常记录（请修正）", padding="5")
        abnormal_frame.pack(fill=BOTH, expand=YES, padx=5, pady=5)
        self.manager_abnormal_records_text = scrolledtext.ScrolledText(abnormal_frame, wrap=tk.WORD, width=80, height=8, font=("微软雅黑", 10))
        self.manager_abnormal_records_text.pack(fill=BOTH, expand=YES)
        
        # 创建右键菜单
        self.manager_abnormal_records_text_menu = tk.Menu(self.manager_abnormal_records_text, tearoff=0)
        self.manager_abnormal_records_text_menu.add_command(label="剪切", command=self.cut_manager_abnormal_records_text)
        self.manager_abnormal_records_text_menu.add_command(label="复制", command=self.copy_manager_abnormal_records_text)
        self.manager_abnormal_records_text_menu.add_command(label="粘贴", command=self.paste_manager_abnormal_records_text)
        self.manager_abnormal_records_text_menu.add_separator()
        self.manager_abnormal_records_text_menu.add_command(label="全选", command=self.select_all_manager_abnormal_records_text)
        
        # 绑定右键事件
        self.manager_abnormal_records_text.bind("<Button-3>", self.show_manager_abnormal_records_text_menu)

        self.status_label = ttk.Label(control_frame, text="请粘贴接龙数据后，点击生成报表按钮。", font=("微软雅黑", 10), bootstyle=INFO)
        self.status_label.pack(pady=10)

        button_frame = ttk.Frame(control_frame)
        button_frame.pack(pady=10)

        self.generate_button = ttk.Button(button_frame, text="生成报表", command=self.start_report_generation, bootstyle=(SUCCESS, OUTLINE))
        self.generate_button.pack(side=LEFT, padx=5, ipadx=10, ipady=5)

        self.clear_button = ttk.Button(button_frame, text="清除数据", command=self.clear_text, bootstyle=(DANGER, OUTLINE))
        self.clear_button.pack(side=LEFT, padx=5, ipadx=10, ipady=5)

    def create_new_business_tab(self, notebook):
        """创建渠道厅店新增业务报表标签页"""
        new_business_tab = ttk.Frame(notebook, padding="10")
        notebook.add(new_business_tab, text="渠道厅店业务")

        # 控制框架（底部）
        control_frame = ttk.Frame(new_business_tab)
        control_frame.pack(side=BOTTOM, fill=X, pady=10)

        # 内容框架（顶部，填充剩余空间）
        content_frame = ttk.Frame(new_business_tab)
        content_frame.pack(side=TOP, fill=BOTH, expand=YES)

        input_frame = ttk.Labelframe(content_frame, text="请粘贴接龙数据", padding="5")
        input_frame.pack(fill=BOTH, expand=YES, padx=5, pady=5)
        self.new_business_text_input = scrolledtext.ScrolledText(input_frame, wrap=tk.WORD, width=80, height=12, font=("微软雅黑", 10))
        self.new_business_text_input.pack(fill=BOTH, expand=YES)
        
        # 创建右键菜单
        self.new_business_text_input_menu = tk.Menu(self.new_business_text_input, tearoff=0)
        self.new_business_text_input_menu.add_command(label="剪切", command=self.cut_new_business_text_input)
        self.new_business_text_input_menu.add_command(label="复制", command=self.copy_new_business_text_input)
        self.new_business_text_input_menu.add_command(label="粘贴", command=self.paste_new_business_text_input)
        self.new_business_text_input_menu.add_separator()
        self.new_business_text_input_menu.add_command(label="全选", command=self.select_all_new_business_text_input)
        
        # 绑定右键事件
        self.new_business_text_input.bind("<Button-3>", self.show_new_business_text_input_menu)

        abnormal_frame = ttk.Labelframe(content_frame, text="异常记录（请修正）", padding="5")
        abnormal_frame.pack(fill=BOTH, expand=YES, padx=5, pady=5)
        self.abnormal_records_text = scrolledtext.ScrolledText(abnormal_frame, wrap=tk.WORD, width=80, height=8, font=("微软雅黑", 10))
        self.abnormal_records_text.pack(fill=BOTH, expand=YES)
        
        # 创建右键菜单
        self.abnormal_records_text_menu = tk.Menu(self.abnormal_records_text, tearoff=0)
        self.abnormal_records_text_menu.add_command(label="剪切", command=self.cut_abnormal_records_text)
        self.abnormal_records_text_menu.add_command(label="复制", command=self.copy_abnormal_records_text)
        self.abnormal_records_text_menu.add_command(label="粘贴", command=self.paste_abnormal_records_text)
        self.abnormal_records_text_menu.add_separator()
        self.abnormal_records_text_menu.add_command(label="全选", command=self.select_all_abnormal_records_text)
        
        # 绑定右键事件
        self.abnormal_records_text.bind("<Button-3>", self.show_abnormal_records_text_menu)

        # 下部：控制按钮和状态

        self.new_business_status_label = ttk.Label(control_frame, text="请粘贴接龙数据后，点击生成报表按钮。", font=("微软雅黑", 10), bootstyle=INFO)
        self.new_business_status_label.pack(pady=10)

        button_frame = ttk.Frame(control_frame)
        button_frame.pack(pady=10)

        self.new_business_generate_button = ttk.Button(button_frame, text="生成报表", command=self.start_new_business_report_generation, bootstyle=(SUCCESS, OUTLINE))
        self.new_business_generate_button.pack(side=LEFT, padx=5, ipadx=10, ipady=5)

        self.new_business_clear_button = ttk.Button(button_frame, text="清除数据", command=self.clear_new_business_text, bootstyle=(DANGER, OUTLINE))
        self.new_business_clear_button.pack(side=LEFT, padx=5, ipadx=10, ipady=5)

    def center_window(self):
        """使窗口在屏幕上居中"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def clear_text(self):
        """清除存量经理报表标签页的文本输入和异常记录"""
        self.text_input.delete("1.0", tk.END)
        self.manager_abnormal_records_text.delete("1.0", tk.END)
        self.status_label.config(text="数据已清除，可以粘贴新的接龙内容。", bootstyle=INFO)
    
    # 右键菜单相关方法
    def show_text_input_menu(self, event):
        """显示存量经理报表文本框的右键菜单"""
        try:
            self.text_input_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.text_input_menu.grab_release()
    
    def cut_text_input(self):
        """剪切存量经理报表文本框中的选中文本"""
        try:
            self.text_input.event_generate("<<Cut>>")
        except tk.TclError:
            pass  # 忽略不支持的操作
    
    def copy_text_input(self):
        """复制存量经理报表文本框中的选中文本"""
        try:
            self.text_input.event_generate("<<Copy>>")
        except tk.TclError:
            pass  # 忽略不支持的操作
    
    def paste_text_input(self):
        """粘贴到存量经理报表文本框"""
        try:
            self.text_input.event_generate("<<Paste>>")
        except tk.TclError:
            pass  # 忽略不支持的操作
    
    def select_all_text_input(self):
        """全选存量经理报表文本框中的文本"""
        self.text_input.tag_add(tk.SEL, "1.0", tk.END)
        self.text_input.mark_set(tk.INSERT, "1.0")
        self.text_input.see(tk.INSERT)
        self.text_input.focus_set()

    def clear_new_business_text(self):
        """清除渠道厅店新增业务文本输入框和异常记录的内容"""
        self.new_business_text_input.delete("1.0", tk.END)
        self.abnormal_records_text.delete("1.0", tk.END)
        self.new_business_status_label.config(text="数据已清除，可以粘贴新的接龙内容。", bootstyle=INFO)
    
    # 通用文本操作方法
    def _show_text_menu(self, menu, event):
        """通用显示文本菜单方法"""
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
    
    def _copy_text(self, text_widget):
        """通用复制文本方法"""
        try:
            text_widget.event_generate("<<Copy>>")
        except tk.TclError:
            pass
    
    def _cut_text(self, text_widget):
        """通用剪切文本方法"""
        try:
            text_widget.event_generate("<<Cut>>")
        except tk.TclError:
            pass
    
    def _paste_text(self, text_widget):
        """通用粘贴文本方法"""
        try:
            text_widget.event_generate("<<Paste>>")
        except tk.TclError:
            pass
    
    def _select_all_text(self, text_widget):
        """通用全选文本方法"""
        text_widget.tag_add(tk.SEL, "1.0", tk.END)
        text_widget.mark_set(tk.INSERT, "1.0")
        text_widget.see(tk.INSERT)
        text_widget.focus_set()
    
    # 存量经理异常记录文本框右键菜单相关方法
    def show_manager_abnormal_records_text_menu(self, event):
        """显示存量经理异常记录文本框的右键菜单"""
        self._show_text_menu(self.manager_abnormal_records_text_menu, event)
    
    def copy_manager_abnormal_records_text(self):
        """复制存量经理异常记录文本框中的选中文本"""
        self._copy_text(self.manager_abnormal_records_text)
    
    def select_all_manager_abnormal_records_text(self):
        """全选存量经理异常记录文本框中的文本"""
        self._select_all_text(self.manager_abnormal_records_text)
    
    def cut_manager_abnormal_records_text(self):
        """剪切存量经理异常记录文本框中的选中文本"""
        self._cut_text(self.manager_abnormal_records_text)
    
    def paste_manager_abnormal_records_text(self):
        """粘贴到存量经理异常记录文本框"""
        self._paste_text(self.manager_abnormal_records_text)
    
    # 渠道厅店新增业务文本框右键菜单相关方法
    def show_new_business_text_input_menu(self, event):
        """显示渠道厅店新增业务文本框的右键菜单"""
        self._show_text_menu(self.new_business_text_input_menu, event)
    
    def cut_new_business_text_input(self):
        """剪切渠道厅店新增业务文本框中的选中文本"""
        self._cut_text(self.new_business_text_input)
    
    def copy_new_business_text_input(self):
        """复制渠道厅店新增业务文本框中的选中文本"""
        self._copy_text(self.new_business_text_input)
    
    def paste_new_business_text_input(self):
        """粘贴到渠道厅店新增业务文本框"""
        self._paste_text(self.new_business_text_input)
    
    def select_all_new_business_text_input(self):
        """全选渠道厅店新增业务文本框中的文本"""
        self._select_all_text(self.new_business_text_input)
    
    # 渠道厅店异常记录文本框右键菜单相关方法
    def show_abnormal_records_text_menu(self, event):
        """显示渠道厅店异常记录文本框的右键菜单"""
        self._show_text_menu(self.abnormal_records_text_menu, event)
    
    def copy_abnormal_records_text(self):
        """复制渠道厅店异常记录文本框中的选中文本"""
        self._copy_text(self.abnormal_records_text)
    
    def select_all_abnormal_records_text(self):
        """全选渠道厅店异常记录文本框中的文本"""
        self._select_all_text(self.abnormal_records_text)
    
    def cut_abnormal_records_text(self):
        """剪切渠道厅店异常记录文本框中的选中文本"""
        self._cut_text(self.abnormal_records_text)
    
    def paste_abnormal_records_text(self):
        """粘贴到渠道厅店异常记录文本框"""
        self._paste_text(self.abnormal_records_text)

    def start_new_business_report_generation(self):
        """开始渠道厅店新增业务报表生成过程"""
        dragon_text = self.new_business_text_input.get("1.0", tk.END).strip()
        if not dragon_text:
            messagebox.showwarning("输入为空", "请先在文本框中粘贴接龙数据！")
            return

        self.new_business_generate_button.config(state=DISABLED)
        self.new_business_status_label.config(text="正在生成报表，请稍候...", bootstyle=DEFAULT)
        
        thread = threading.Thread(target=generate_new_business_report, args=(dragon_text, self.update_new_business_status, self.config_dir))
        thread.daemon = True
        thread.start()

    def start_report_generation(self):
        """开始报表生成过程，使用线程以避免GUI冻结"""
        dragon_text = self.text_input.get("1.0", tk.END).strip()
        if not dragon_text:
            messagebox.showwarning("输入为空", "请先在文本框中粘贴接龙数据！")
            return

        self.generate_button.config(state=DISABLED)
        self.status_label.config(text="正在生成报表，请稍候...", bootstyle=DEFAULT)
        
        # 在新线程中运行报表生成函数
        thread = threading.Thread(target=generate_report, args=(dragon_text, self.update_manager_status, self.config_dir))
        thread.daemon = True
        thread.start()

    def update_new_business_status(self, message, style, abnormal_text=None):
        """线程安全地更新渠道厅店新增业务GUI状态"""
        def _update():
            self.new_business_status_label.config(text=message, bootstyle=style.upper())
            self.new_business_generate_button.config(state=NORMAL)
            if abnormal_text is not None:
                self.abnormal_records_text.delete("1.0", tk.END)
                self.abnormal_records_text.insert(tk.END, abnormal_text)
            else:
                self.abnormal_records_text.delete("1.0", tk.END) # 如果没有异常信息则清空
        self.after(0, _update)

    def update_manager_status(self, message, style, abnormal_text=None):
        """线程安全地更新存量经理报表GUI状态"""
        def _update():
            self.status_label.config(text=message, bootstyle=style.upper())
            self.generate_button.config(state=NORMAL)
            if abnormal_text is not None:
                self.manager_abnormal_records_text.delete("1.0", tk.END)
                self.manager_abnormal_records_text.insert(tk.END, abnormal_text)
            else:
                self.manager_abnormal_records_text.delete("1.0", tk.END)
        self.after(0, _update)

def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='通用接龙数据报表生成器')
    parser.add_argument('--config-dir', type=str, help='配置文件目录路径')
    parser.add_argument('--gui', action='store_true', help='启动图形界面（默认）')
    parser.add_argument('--no-gui', action='store_true', help='不启动图形界面')
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_args()
    
    # 设置全局配置目录
    global_config_dir = args.config_dir
    
    # 更新配置管理器
    if global_config_dir:
        config_manager = ConfigManager(global_config_dir)
    
    if args.no_gui:
        print("命令行模式已启用，请使用图形界面模式")
        sys.exit(1)
    else:
        # 启动GUI
        app = ReportApp()
        app.mainloop()